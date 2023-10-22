# 确保已安装必要的第三方库
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


options = Options()
# options.add_argument('--headless')
# options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
# 指定浏览器为Chrome，注意需先把IEDriverServer.exe放在python安装目录
driver = webdriver.Chrome(service=Service(
    ChromeDriverManager().install()), options=options)

# 打开姓名excel表
wb = openpyxl.load_workbook('/home/tz/Documents/Web_Crawler/Excel/Test.xlsx')
# 将活动工作表设置赋值给ws
ws = wb.active
# 获取表格行数
r = ws.max_row


# 打开要填表的网页链接。
driver.get(r'https://www.xingming.com/')
time.sleep(2)

# 窗口最大化
driver.maximize_window()

# 将excel各内容行循环填入网页，并获取查询结果中的指定内容另存至excel

for i in range(2, r + 1):

    # 先清空输入框内容
    driver.find_element_by_name('xs').clear()
    driver.find_element_by_name('mz').clear()

    # 输入姓氏
    driver.find_element_by_name('xs').send_keys(ws.cell(i, 2).value)

    # 输入名字
    driver.find_element_by_name('mz').send_keys(ws.cell(i, 3).value)

    # 点提交按钮
    driver.find_element_by_xpath(
        '/html/body/div[4]/div[1]/div[2]/form/div/input[3]').click()

    # 等待1秒待网页加载完毕，可根据网速或停留需求自定义时长
    time.sleep(1)

    # 获取评分段落对应文字，填入excel对应名字后方单元格
    result = driver.find_element_by_xpath('/html/body/div[4]/div[1]/div/p[4]')
    ws.cell(i, 4).value = result.text

    # 跳回姓名查询页面
    if i < r:
        driver.back()

# 为新增列设置标题
ws.cell(1, 4).value = '评分'

# 将"姓名.xlsx"另存为"姓名评分.xlsx"
wb.save('/home/tz/Documents/Web_Crawler/Excel/Output/test_result.xlsx')
