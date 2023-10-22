import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


options = Options()
options.add_argument('--disable-dev-shm-usage')
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# open excel
wb = openpyxl.load_workbook('/home/tz/Documents/Web_Crawler/Excel/avis.xlsx')
ws = wb.active
r = ws.max_row


# open link。
driver.get(r'https://www.avis.de/business/firmenkunden/fahrzeug-bruttolistenpreis#')

# maxmal windows
driver.maximize_window()

driver.find_element(by=By.XPATH, value='/html/body/div[4]/div/div/div/div[1]/a[3]').click()

# find the iframes
iframes = driver.find_elements(By.TAG_NAME, 'iframe')
print(f"Found {len(iframes)} iframes on the page.")
for index, iframe in enumerate(iframes):
    print(f"Iframe {index}: {iframe.get_attribute('src')}")

driver.switch_to.frame(0)
# driver.switch_to.default_content()

driver.find_element(By.NAME, 'KENNZ_0').clear()

WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, 'KENNZ_0')))

for i in range(2, r + 1):

    driver.find_element(By.NAME, 'KENNZ_0').clear()

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.NAME, 'KENNZ_0'))
    ).send_keys(ws.cell(i, 1).value)
    

    driver.find_element(by=By.XPATH, value='/html/body/form/div/div[2]/table/tbody/tr/td/div[2]/p/input[1]').click()
    time.sleep(0.1)

    # Modell
    result2 = driver.find_element(by=By.XPATH, value=
        '/html/body/form/div/div[2]/table/tbody/tr/td/div[1]/div/table/tbody/tr[2]/td[2]/div')
    
    # Fahrgestellnummer
    result3 = driver.find_element(by=By.XPATH, value=
        '/html/body/form/div/div[2]/table/tbody/tr/td/div[1]/div/table/tbody/tr[2]/td[3]/div')
    
    # Bruttolistenpreis
    result4 = driver.find_element(by=By.XPATH, value=
        '/html/body/form/div/div[2]/table/tbody/tr/td/div[1]/div/table/tbody/tr[2]/td[4]/div')

    # KW
    result5 = driver.find_element(by=By.XPATH, value=
        '/html/body/form/div/div[2]/table/tbody/tr/td/div[1]/div/table/tbody/tr[2]/td[5]/div')
    
    # CO2
    result6 = driver.find_element(by=By.XPATH, value=
        '/html/body/form/div/div[2]/table/tbody/tr/td/div[1]/div/table/tbody/tr[2]/td[6]/div')

    ws.cell(i, 2).value = result2.text
    ws.cell(i, 3).value = result3.text
    ws.cell(i, 4).value = result4.text
    ws.cell(i, 5).value = result5.text
    ws.cell(i, 6).value = result6.text

driver.switch_to.default_content()

ws.cell(1, 2).value = 'Modell'
ws.cell(1, 3).value = 'Fahrgestellnummer'
ws.cell(1, 4).value = 'Bruttolistenpreis'
ws.cell(1, 5).value = 'KW'
ws.cell(1, 6).value = 'CO2'


# save"avis.xlsx"under save"avis_result.xlsx"
wb.save('/home/tz/Documents/Web_Crawler/Excel/Output/avis_result.xlsx')